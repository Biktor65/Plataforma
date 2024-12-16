from flask import Blueprint, jsonify, render_template, request, redirect, url_for, session
from .models import fetch_data_from_sql
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload
from app.usuarios.views import login_required, role_required
import io
import openpyxl
import json
import os 
from dotenv import load_dotenv

load_dotenv()

clientes_bp = Blueprint('clientes', __name__)

CARPETA_ID = os.getenv('CARPETA_ID')
FILE_ID = os.getenv('FILE_ID')
SERVICE_ACCOUNT_FILE = os.getenv('SERVICE_ACCOUNT_FILE')

credentials = service_account.Credentials.from_service_account_file(
    SERVICE_ACCOUNT_FILE,
    scopes=["https://www.googleapis.com/auth/drive"]
)

drive_service = build('drive', 'v3', credentials=credentials)

@clientes_bp.route('/intro_envase')
@login_required
@role_required('usuario')
def intro_envase():
    return render_template('intro_envase.html')

@clientes_bp.route('/api/clientes', methods=['GET'])
@login_required
@role_required('usuario')
def get_clientes():
    data = fetch_data_from_sql()
    if isinstance(data, str):  # Si hay un error
        return jsonify({"error": data}), 500
    return jsonify(data)

@clientes_bp.route('/enviar_formulario', methods=['POST'])
@login_required
@role_required('usuario')
def enviar_formulario():
    # Descargar el archivo de Google Drive
    request_drive = drive_service.files().get_media(fileId=FILE_ID)
    file_stream = io.BytesIO()
    downloader = MediaIoBaseDownload(file_stream, request_drive)
    done = False
    while not done:
        status, done = downloader.next_chunk()

    # Cargar el archivo en memoria usando openpyxl
    file_stream.seek(0)
    workbook = openpyxl.load_workbook(file_stream)
    sheet = workbook.active

    new_row_index = None
    for row in range(1, sheet.max_row + 1): 
        if sheet.cell(row=row, column=1).value is None:
            new_row_index = row
            break
    if new_row_index is None:
        new_row_index = sheet.max_row + 1        
        
    # Extraer datos del formulario
    data = request.form
    cliente_info = json.loads(data['CODCliente']) if isinstance(data['CODCliente'], str) and data['CODCliente'].startswith('{') else {}

    # Crear la nueva fila con el formato correcto
    nueva_fila = [
        cliente_info.get("CODCliente", data.get('CODCliente', '')),
        data.get('NumeroIdentidadCliente', ''),
        data.get('NombreComercial', ''),
        data.get('DireccionNegocio', ''),
        data.get('Motivo', ''),
        data.get('Beneficio', ''),
        data.get('Fecha', ''),
        data.get('Introduccion', ''),
        data.get('Ruta', ''),
        data.get('Pepsi6_5oz_cant', ''),
        data.get('Pepsi6_5oz_precio', ''),
        data.get('Pepsi12oz_cant', ''),
        data.get('Pepsi12oz_precio', ''),
        data.get('Pepsi500ml_cant', ''),
        data.get('Pepsi500ml_precio', ''),
        data.get('Pepsi1_25lts_cant', ''),
        data.get('Pepsi1_25lts_precio', ''),
        data.get('MirindaN12oz_cant', ''),
        data.get('MirindaN12oz_precio', ''),
        data.get('MirindaB12oz_cant', ''),
        data.get('MirindaB12oz_precio', ''),
        data.get('MirindaU12oz_cant', ''),
        data.get('MirindaU12oz_precio', ''),
        data.get('MirindaN1_25lts_cant', ''),
        data.get('MirindaN1_25lts_precio', ''),
        data.get('sevenup12oz_cant', ''),
        data.get('sevenup12oz_precio', ''),
        data.get('teem12oz_cant', ''),
        data.get('teem12oz_precio', ''),
        data.get('LinkM6_5oz_cant', ''),
        data.get('LinkM6_5oz_precio', ''),
        data.get('LinkB6_5oz_cant', ''),
        data.get('LinkB6_5oz_precio', ''),
        data.get('LinkC6_5oz_cant', ''),
        data.get('LinkC6_5oz_precio', ''),
    ]

    # Agregar la nueva fila y guardar localmente
    for col_index, value in enumerate(nueva_fila, start=1):
        sheet.cell(row=new_row_index, column=col_index, value=value)
        
    # Subir el archivo actualizado directamente desde la memoria a Google Drive
    output_stream = io.BytesIO()
    workbook.save(output_stream)
    output_stream.seek(0)
    media = MediaIoBaseUpload(output_stream, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    drive_service.files().update(fileId=FILE_ID, media_body=media).execute()
    print("Archivo subido exitosamente a Google Drive desde la memoria.")

    return "Datos enviados y archivo actualizado en Google Drive."
